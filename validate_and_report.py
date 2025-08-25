import json
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
from pathlib import Path

from jsonschema import Draft7Validator
from openpyxl import Workbook
from datetime import datetime
import re
from pypdf import PdfReader


@dataclass
class Counts:
    toc_total: int
    spec_total: int
    missing_in_spec: int
    extra_in_spec: int
    order_errors: int


@dataclass
class ValidationResult:
    counts: Counts
    missing: List[str]
    extra: List[str]
    order_mismatches: List[str]
    toc_gaps: List[Tuple[str, str]]
    spec_gaps: List[Tuple[str, str]]
    errors_toc: List[Tuple[int, str]]
    errors_spec: List[Tuple[int, str]]
    toc_table_ids: List[str]
    doc_table_ids: List[str]


def read_jsonl(path: str) -> List[dict]:
    """Read JSONL file and return list of dictionaries."""
    rows: List[dict] = []
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"Error reading {path}: {e}")
        return []
    except (FileNotFoundError, 
            json.JSONDecodeError) as e:
        print(f"Error reading {path}: {e}")
        return []
    return rows


def load_schema(schema_path: str) -> Optional[dict]:
    """Load JSON schema file."""
    try:
        with open(schema_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"Error loading schema {schema_path}: {e}")
        return None


def validate_rows(rows: List[dict], schema_path: str) -> List[Tuple[int, str]]:
    """Validate rows against JSON schema."""
    schema = load_schema(schema_path)
    if not schema:
        return []
    
    validator = Draft7Validator(schema)
    errors: List[Tuple[int, str]] = []
    
    for idx, row in enumerate(rows):
        for error in validator.iter_errors(row):
            errors.append((idx, error.message))
    
    return errors


def get_section_ids(rows: List[dict]) -> List[str]:
    """Extract section IDs from rows."""
    return [row["section_id"] for row in rows]


def find_missing_and_extra_sections(
    toc_ids: List[str], spec_ids: List[str]
) -> Tuple[List[str], List[str]]:
    """Find missing and extra sections between TOC and spec."""
    toc_set = set(toc_ids)
    spec_set = set(spec_ids)
    
    missing = sorted(list(toc_set - spec_set))
    extra = sorted(list(spec_set - toc_set))
    
    return missing, extra


def find_order_mismatches(toc_ids: List[str], spec_ids: List[str]) -> List[str]:
    """Find order mismatches between TOC and spec."""
    spec_order_index = {sid: i for i, sid in enumerate(spec_ids)}
    order_mismatches = []
    last_index = -1
    
    for sid in toc_ids:
        if sid in spec_order_index:
            idx = spec_order_index[sid]
            if idx < last_index:
                order_mismatches.append(sid)
            last_index = idx
    
    return order_mismatches


def compare_toc_vs_spec(
    toc_rows: List[dict], spec_rows: List[dict]
) -> Tuple[Counts, List[str], List[str], List[str]]:
    """Compare TOC vs spec and return counts and differences."""
    toc_ids = get_section_ids(toc_rows)
    spec_ids = get_section_ids(spec_rows)
    
    missing, extra = find_missing_and_extra_sections(toc_ids, spec_ids)
    order_mismatches = find_order_mismatches(toc_ids, spec_ids)
    
    counts = Counts(
        toc_total=len(toc_ids),
        spec_total=len(spec_ids),
        missing_in_spec=len(missing),
        extra_in_spec=len(extra),
        order_errors=len(order_mismatches),
    )
    
    return counts, missing, extra, order_mismatches


def group_sections_by_parent(section_ids: List[str]) -> Dict[str, List[str]]:
    """Group section IDs by their parent."""
    by_parent: Dict[str, List[str]] = {}
    
    for sid in section_ids:
        parent = (".".join(sid.split(".")[:-1]) 
                 if "." in sid else "<root>")
        by_parent.setdefault(parent, []).append(sid)
    
    return by_parent


def get_numeric_last_component(section_id: str) -> int:
    """Extract the numeric last component of a section ID."""
    try:
        return int(section_id.split(".")[-1])
    except (ValueError, IndexError):
        return -1


def find_missing_numbers_in_sequence(nums: List[int]) -> List[int]:
    """Find missing numbers in a sequence."""
    if not nums:
        return []
    
    expected = list(range(nums[0], nums[-1] + 1))
    return [n for n in expected if n not in nums]


def process_parent_group(parent: str, children: List[str]) -> Optional[Tuple[str, str]]:
    """Process a single parent group to find gaps."""
    if len(children) < 2:
        return None
    
    # Only consider same-depth children
    depth = len(children[0].split(".")) if children else 0
    same_depth = [c for c in children if len(c.split(".")) == depth]
    
    if len(same_depth) < 2:
        return None
    
    # Sort by numeric last component
    same_depth.sort(key=get_numeric_last_component)
    nums = [
        get_numeric_last_component(s) for s in same_depth 
        if get_numeric_last_component(s) >= 0
    ]
    nums = [
        get_numeric_last_component(s) for s in same_depth 
        if get_numeric_last_component(s) >= 0
    ]
    
    if not nums:
        return None
    
    missing_nums = find_missing_numbers_in_sequence(nums)
    if not missing_nums:
        return None
    
    missing_ids = [
        f"{parent}.{n}" if parent != "<root>" else str(n) 
        for n in missing_nums
    ]
    missing_ids = [
        f"{parent}.{n}" if parent != "<root>" else str(n) 
        for n in missing_nums
    ]
    
    parent_label = parent if parent != "<root>" else "<top-level>"
    gap_description = ", ".join(missing_ids)
    gap_description = ", ".join(missing_ids)
    
    return (parent_label, gap_description)


def detect_gaps(section_ids: List[str]) -> List[Tuple[str, str]]:
    """Detect numeric gaps among siblings."""
    by_parent = group_sections_by_parent(section_ids)
    gaps: List[Tuple[str, str]] = []
    
    for parent, children in by_parent.items():
        gap_info = process_parent_group(parent, children)
        if gap_info:
            gaps.append(gap_info)
    
    return gaps


def extract_list_of_tables_text(reader: PdfReader, max_scan_pages: int = 120) -> str:
    """Extract text from the beginning of the PDF for table list."""
    text_parts: List[str] = []
    
    for i in range(min(max_scan_pages, len(reader.pages))):
        text = reader.pages[i].extract_text()
        if text:
            text_parts.append(text)
    
    return "\n".join(text_parts)


def parse_tables_from_list(front_text: str) -> List[str]:
    """Parse 'List of Tables' style entries into table IDs."""
    ids: List[str] = []
    pattern = re.compile(
        r"^Table\s+(\d+-\d+)[^\n]*?(?:\s+\d+)?\s*$", 
        re.MULTILINE
    )
    
    for match in pattern.finditer(front_text):
        ids.append(match.group(1))
    
    # Deduplicate preserving order
    seen = set()
    unique_ids: List[str] = []
    
    for table_id in ids:
        if table_id not in seen:
            seen.add(table_id)
            unique_ids.append(table_id)
    
    return unique_ids


def scan_tables_in_document(reader: PdfReader) -> List[str]:
    """Scan entire PDF text for 'Table X-Y' labels and return unique IDs."""
    found: List[str] = []
    pattern = re.compile(
        r"^Table\s+(\d+-\d+)[^\n]*$", 
        re.MULTILINE
    )
    
    for page in reader.pages:
        text = page.extract_text()
        if text:
            for match in pattern.finditer(text):
                found.append(match.group(1))
    
    # Deduplicate preserving order
    seen = set()
    unique_ids: List[str] = []
    
    for table_id in found:
        if table_id not in seen:
            seen.add(table_id)
            unique_ids.append(table_id)
    
    return unique_ids


def create_overview_sheet(wb: Workbook, counts: Counts, errors_toc: List, errors_spec: List) -> None:
    """Create the overview sheet with summary statistics."""
    ws_overview = wb.active
    ws_overview.title = "overview"
    
    headers = [
        "toc_total", "spec_total", "missing_in_spec", "extra_in_spec",
        "order_errors", "toc_schema_errors", "spec_schema_errors"
    ]
    values = [
        counts.toc_total, counts.spec_total, counts.missing_in_spec,
        counts.extra_in_spec, counts.order_errors, len(errors_toc), len(errors_spec)
    ]
    
    ws_overview.append(headers)
    ws_overview.append(values)


def create_analysis_sheets(
    wb: Workbook, missing: List[str], extra: List[str], 
    order_mismatches: List[str], toc_gaps: List[Tuple], 
    spec_gaps: List[Tuple]
) -> None:
    """Create analysis sheets for different types of issues."""
    # Missing sections
    ws_missing = wb.create_sheet("missing_in_spec")
    ws_missing.append(["missing_in_spec"])
    for section_id in missing:
        ws_missing.append([section_id])

    # Extra sections
    ws_extra = wb.create_sheet("extra_in_spec")
    ws_extra.append(["extra_in_spec"])
    for section_id in extra:
        ws_extra.append([section_id])

    # Order mismatches
    ws_order = wb.create_sheet("order_mismatches")
    ws_order.append(["order_mismatches"])
    for section_id in order_mismatches:
        ws_order.append([section_id])

    # TOC gaps
    ws_gaps_toc = wb.create_sheet("gaps_toc")
    ws_gaps_toc.append(["parent_id", "missing_children"])
    for parent, missing_list in toc_gaps:
        ws_gaps_toc.append([parent, missing_list])

    # Spec gaps
    ws_gaps_spec = wb.create_sheet("gaps_spec")
    ws_gaps_spec.append(["parent_id", "missing_children"])
    for parent, missing_list in spec_gaps:
        ws_gaps_spec.append([parent, missing_list])


def create_error_sheets(wb: Workbook, errors_toc: List, errors_spec: List) -> None:
    """Create sheets for schema validation errors."""
    if errors_toc:
        ws_toc_err = wb.create_sheet("toc_schema_errors")
        ws_toc_err.append(["row_index", "error"]) 
        for idx, msg in errors_toc:
            ws_toc_err.append([idx, msg])
    
    if errors_spec:
        ws_spec_err = wb.create_sheet("spec_schema_errors")
        ws_spec_err.append(["row_index", "error"]) 
        for idx, msg in errors_spec:
            ws_spec_err.append([idx, msg])


def create_table_analysis_sheet(
    wb: Workbook, toc_table_ids: List[str], doc_table_ids: List[str]
) -> None:
    """Create sheet for table count analysis."""
    ws_tables = wb.create_sheet("tables")
    ws_tables.append(["metric", "value"]) 
    ws_tables.append(["toc_tables_count", len(toc_table_ids)])
    ws_tables.append(["parsed_tables_count", len(doc_table_ids)])
    
    # Differences
    toc_table_set = set(toc_table_ids)
    doc_table_set = set(doc_table_ids)
    missing_tables = [t for t in toc_table_ids if t not in doc_table_set]
    extra_tables = [t for t in doc_table_ids if t not in toc_table_set]
    
    ws_tables.append(["missing_tables", ", ".join(missing_tables)])
    ws_tables.append(["extra_tables", ", ".join(extra_tables)])


def save_workbook_with_fallback(workbook: Workbook, primary_path: str) -> str:
    """Save workbook with fallback to timestamped filename."""
    try:
        workbook.save(primary_path)
        return primary_path
    except PermissionError:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        alt_path = f"validation_report_{timestamp}.xlsx"
        workbook.save(alt_path)
        return alt_path


def perform_validation() -> ValidationResult:
    """Perform all validation steps and return results."""
    print("📖 Reading TOC file...")
    toc_rows = read_jsonl("usb_pd_toc.jsonl")
    print(f"   Read {len(toc_rows)} TOC entries")
    
    print("📖 Reading sections file...")
    spec_rows = read_jsonl("usb_pd_spec.jsonl")
    print(f"   Read {len(spec_rows)} section entries")

    print("🔍 Validating TOC schema...")
    errors_toc = validate_rows(toc_rows, "schema_toc.json")
    print(f"   Found {len(errors_toc)} TOC schema errors")
    
    print("🔍 Validating sections schema...")
    errors_spec = validate_rows(spec_rows, "schema_sections.json")
    print(f"   Found {len(errors_spec)} sections schema errors")

    print("⚖️ Comparing TOC vs sections...")
    counts, missing, extra, order_mismatches = compare_toc_vs_spec(toc_rows, spec_rows)
    print(f"   TOC: {counts.toc_total}, Sections: {counts.spec_total}")
    print(f"   Missing: {counts.missing_in_spec}, Extra: {counts.extra_in_spec}")

    print("🔍 Detecting gaps in TOC...")
    toc_gaps = detect_gaps([r["section_id"] for r in toc_rows])
    print(f"   Found {len(toc_gaps)} TOC gaps")
    
    print("🔍 Detecting gaps in sections...")
    spec_gaps = detect_gaps([r["section_id"] for r in spec_rows])
    print(f"   Found {len(spec_gaps)} section gaps")

    print("📋 Analyzing tables...")
    reader = PdfReader("usb_pd_spec.pdf")
    front_text = extract_list_of_tables_text(reader)
    toc_table_ids = parse_tables_from_list(front_text)
    doc_table_ids = scan_tables_in_document(reader)
    print(f"   TOC tables: {len(toc_table_ids)}, Document tables: {len(doc_table_ids)}")

    return ValidationResult(
        counts=counts,
        missing=missing,
        extra=extra,
        order_mismatches=order_mismatches,
        toc_gaps=toc_gaps,
        spec_gaps=spec_gaps,
        errors_toc=errors_toc,
        errors_spec=errors_spec,
        toc_table_ids=toc_table_ids,
        doc_table_ids=doc_table_ids
    )


def create_excel_report(result: ValidationResult) -> str:
    """Create Excel report from validation results."""
    print("📊 Creating Excel workbook...")
    wb = Workbook()
    
    print("📋 Creating overview sheet...")
    create_overview_sheet(wb, result.counts, result.errors_toc, result.errors_spec)
    
    print("📋 Creating analysis sheets...")
    create_analysis_sheets(
        wb, result.missing, result.extra, result.order_mismatches,
        result.toc_gaps, result.spec_gaps
    )
    
    print("📋 Creating error sheets...")
    create_error_sheets(wb, result.errors_toc, result.errors_spec)
    
    print("📋 Creating table analysis sheet...")
    create_table_analysis_sheet(wb, result.toc_table_ids, result.doc_table_ids)

    print("💾 Saving Excel file...")
    output_path = save_workbook_with_fallback(wb, "validation_report.xlsx")
    print(f"✅ Wrote {output_path}")
    
    return output_path


def main() -> None:
    """Main function to run the validation and generate report."""
    print("📊 Starting validation and report generation...")
    
    try:
        # Perform validation
        result = perform_validation()
        
        # Create report
        create_excel_report(result)
        
        print("🎉 Validation report complete!")
        
    except Exception as e:
        print(f"❌ Error during validation: {e}")
        raise


if __name__ == "__main__":
    main()

